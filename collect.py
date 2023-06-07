#!/usr/bin/python3

# https://discovery.googleapis.com/discovery/v1/apis
# https://developers.google.com/discovery/v1/reference?hl=en

import argparse
import csv
import json
import os
import subprocess
import datetime
import pandas as pd
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from google.cloud import compute_v1
from googleapiclient import discovery
from google.cloud import monitoring_v3
from google.auth import default as gcp_cred
from easydict import EasyDict
from collections import defaultdict

warnings.filterwarnings("ignore", "Your application has authenticated using end user credentials")

class Logger:
    level = 1  # Default log level

    @classmethod
    def set_level(cls, level):
        cls.level = level

    @classmethod
    def log(cls, level, message):
        if cls.level >= level:
            print(message)

class Exporter:
    # Save the desired outputs
    outputs = []
    out_dir = "output"

    # To make date consistent between the outputs
    today = datetime.datetime.today()
    date = today.strftime('%d/%m/%y %H:%M:%S')
 
    @classmethod
    def set_outputs(cls, outputs):
        cls.outputs = outputs
        os.makedirs(cls.out_dir, exist_ok=True)

        # Delete the file if it exists
        if 'xls' in outputs:
            xls_path = os.path.join(cls.out_dir, 'output.xlsx')
            if os.path.exists(xls_path):
                os.remove(xls_path)

    @classmethod
    def export(cls, data, resource_type):
        Logger.log(1, f' | - Saving {resource_type}')

        if len(data) == 0:
            return

        if 'csv' in cls.outputs:
            cls.save_to_csv(data, resource_type)
        if 'xls' in cls.outputs:
            cls.save_to_xls(data, resource_type)
        if 'json' in cls.outputs:
            cls.save_to_json(data, resource_type)
    
    @classmethod
    def save_to_csv(cls, data, resource_type):
        output_path = os.path.join(cls.out_dir, f"output_{resource_type}.csv")
        with open(output_path, 'w', newline='') as csvfile:
            # Write a line (commented) that has some metadata info about this file
            csvfile.write(f'# Type: {resource_type} Date: {cls.date}\n')
            # Dump the Dict variable
            writer = csv.DictWriter(csvfile, fieldnames=data[0].keys())
            # Write the header
            writer.writeheader()
            # Plot the data
            writer.writerows(data)

    @classmethod
    def save_to_json(cls, data, resource_type):
        json_path = os.path.join(cls.out_dir, f'output_{resource_type}.json')
        with open(json_path, 'w') as file:
            json.dump(data, file, indent=4)

    @classmethod
    def save_to_xls(cls, data, resource_type):
        xls_path = os.path.join(cls.out_dir, 'output.xlsx')
        
        # Check if the file exists
        file_exists = os.path.exists(xls_path)
        
        # If the file exists, load it into the workbook object
        if file_exists:
            workbook = load_workbook(xls_path)
        else:
            workbook = Workbook()
            # Remove the default sheet created when creating a new workbook
            # Change to have a first sheet as a resume to all the data collected
            default_sheet = workbook.active
            workbook.remove(default_sheet)

        # Create a new sheet or get the existing sheet by name
        if resource_type in workbook.sheetnames:
            sheet = workbook[resource_type]
            sheet.delete_rows(1, sheet.max_row)  # Clear existing content
        else:
            sheet = workbook.create_sheet(resource_type)

        # Convert data to a DataFrame
        df = pd.DataFrame(data)

        # Append the DataFrame to the sheet
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)

        # Apply bold formatting to the first row
        first_row = sheet[1]
        for cell in first_row:
            cell.font = Font(bold=True)

        # Adjust column width to fit the content
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            # Look for the cell text size, using the fixed value 40 as max
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 40)
                except TypeError:
                    pass
            # Put 20% more to make a safe width visualization
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Apply color scheme to the table
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        header_row = sheet[1]
        for cell in header_row:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = data_fill

        # Save the workbook
        workbook.save(xls_path)

def GetHumanReadable(size,precision=2):
    suffixes=['B','KB','MB','GB','TB']
    suffixIndex = 0
    while size > 1024 and suffixIndex < 4:
        suffixIndex += 1 #increment the index of the suffix
        size = size/1024.0 #apply the division
    return "%.*f%s"%(precision,size,suffixes[suffixIndex])

def check_api_is_enabled(project_id, api):
    service = discovery.build('serviceusage', 'v1')
    request = service.services().get(
        name=f"projects/{project_id}/services/{api}.googleapis.com"
    )
    response = request.execute()

    return True if response.get('state') != "DISABLED" else False

def get_logged_in_project():
    credentials, project_id = gcp_cred()
    return project_id

def list_vm_instances(project_id, gke_ignore):
    def get_vm_metrics(vm_name):
        # Set the time range for the metrics query (last month)
        end_time = datetime.datetime.utcnow()
        start_time = end_time - datetime.timedelta(days=30)

        interval = monitoring_v3.TimeInterval(
            {
                "start_time":{"seconds": int(start_time.timestamp())},
                "end_time":{"seconds": int(end_time.timestamp())}
            }
        )

        # Define the resource for the VM instance
        resource = f"projects/{project_id}/zones/{zone}/instances/{vm_name}"

        metric_filter=f'metric.type = "compute.googleapis.com/instance/cpu/utilization" AND resource.type="gce_instance" AND metric.label.instance_name="{vm_name}"'

        #print(metric_filter)

        # Prepare the request
        monitoring_client = monitoring_v3.MetricServiceClient()
        response = monitoring_client.list_time_series(
            request={
                "name": f'projects/{project_id}',
                "filter": metric_filter,
                "interval": interval,
                "view": monitoring_v3.ListTimeSeriesRequest.TimeSeriesView.FULL,
            }
        )

        # Extract the data points
        data_points = [point.value.double_value for series in response.time_series for point in series.points]

        # Store the metric values
        results = {}
        results['compute.googleapis.com/instance/cpu/utilization'] = data_points

        return results

    def calculate_average_usage(metric_data):
        # Calculate the average CPU and memory usage
        cpu_usage = metric_data['compute.googleapis.com/instance/cpu/utilization']
        avg_cpu_usage = sum(cpu_usage) / len(cpu_usage) if cpu_usage else 0
        return avg_cpu_usage

    def calculate_p95_usage(metric_data):
        cpu_usage = metric_data['compute.googleapis.com/instance/cpu/utilization']
        p95_cpu_usage = sorted(cpu_usage)[int(len(cpu_usage) * 0.95)] if cpu_usage else 0
        return p95_cpu_usage

    # -----------
    # Start
    Logger.log (1, f"PRJ: {project_id} - Compute Machines")

    instances = []

    if not check_api_is_enabled(project_id,'compute'):
        Logger.log(2, "  | Compute API is not enabled")
        return instances

    compute_client = compute_v1.InstancesClient()
    machineClient = compute_v1.MachineTypesClient()

    request = compute_v1.AggregatedListInstancesRequest()
    request.project = project_id
    # Use the `max_results` parameter to limit the number of results that the API returns per response page.
    request.max_results = 50

    agg_list = compute_client.aggregated_list(request=request)

    #all_instances = defaultdict(list)
    for zone, response in agg_list:
        if response.instances:
            #all_instances[zone].extend(response.instances)
            Logger.log (2, f" |- Found {len(response.instances)} in zone {zone}")
            total=len(response.instances)
            i=0
            end='\r'

            for instance in response.instances:
                print(f'Reading: {i+1}/{total} Zone: {zone} Name: {instance.name}', end=f'{end}')
                i+=1
                if i==total:
                    print()

                # Check for the goog-gke-node label
                IS_GKE=True if 'goog-gke-node' in instance.labels else False

                if gke_ignore:
                    if IS_GKE:
                        Logger.log(3, 'Compute is a GKE Node. Next')
                        continue

                # Get the metrics of the vm, to calculate avg and p95 cpu usage
                metric_data = get_vm_metrics(instance.name)
                avg_cpu_usage = calculate_average_usage(metric_data)
                p95_cpu_usage = calculate_p95_usage(metric_data)

                Logger.log (3, f"Avg: {avg_cpu_usage} - p95: {p95_cpu_usage}")

                parts=instance.machine_type.split('/')
                machine_type=parts[-1]

                # Get Size and Type of OS
                diskSize=0
                SO="NOT Win"
                for disk in instance.disks:
                    diskSize += disk.disk_size_gb
                    contains_windows = any("WINDOWS" in feature.type for feature in disk.guest_os_features)
                    if contains_windows:
                        SO="Windows"

                # Check if preemptible
                PREEMPTIBLE=False
                if 'scheduling' in instance:
                    PREEMPTIBLE=instance.scheduling.preemptible

                # Find the type of the machinetype
                machine=machineClient.get(
                        project=project_id,
                        zone=zone.split('/')[-1],
                        machine_type=machine_type
                    )

                instance_data = {
                    'project': project_id,
                    'name': instance.name,
                    'zone': zone.split('/')[-1],
                    'status': instance.status,
                    'machine_type': machine_type,
                    'Cpu': machine.guest_cpus,
                    'MemMb': machine.memory_mb,
                    'diskSizeGb': diskSize,
                    'SO': SO,
                    'IsGke': IS_GKE,
                    'Preemptible': PREEMPTIBLE,
                    'labels': instance.labels,
                    'cpu_avg': avg_cpu_usage,
                    'cpu_p95': p95_cpu_usage
                }
                instances.append(instance_data)

    return instances

def list_cloudsql_instances(project_id):
    Logger.log (1, f"PRJ: {project_id} - CloudSQL")

    # -----
    # https://stackoverflow.com/questions/69658130/how-can-i-get-list-of-all-cloud-sql-gcp-instances-which-are-stopped-in-pytho
    instances = []

    # Look for the instances in this project
    sql_client = discovery.build('sqladmin', 'v1beta4')
    raw_resp = sql_client.instances().list(project=project_id).execute()
    resp = EasyDict(raw_resp)

    if not 'items' in resp:
        return instances

    Logger.log(2, f' |- Found {len(resp.items)}')

    for inst in resp.items:
        instance_data = {
            'project': project_id,
            'name': inst.name,
            'database_version': inst.databaseVersion,
            'installedVersion': inst.databaseInstalledVersion,
            'tier': inst.settings.tier,
            'instanceType': inst.instanceType,
            'region': inst.region,
            'diskType': inst.settings.dataDiskType,
            'diskSizeGb': inst.settings.dataDiskSizeGb
        }
        instances.append(instance_data)

    return instances

def list_functions(project_id):
    Logger.log (1, f"PRJ: {project_id} - Functions")

    functions = []

    if not check_api_is_enabled(project_id,'cloudfunctions'):
        Logger.log (2, "  | cloudfunctions API not enabled")
        return functions

    client = discovery.build('cloudfunctions', 'v1')
    # Note the use of - (dash) indicating all locations
    raw_resp = client.projects().locations().functions().list(
            parent=f"projects/{project_id}/locations/-"
        ).execute()
    resp = EasyDict(raw_resp)

    Logger.log(2, f' |- Found {len(resp.functions)}')

    for fnc in resp.functions:
        #print(f'|{fnc}|')
        parts = fnc.name.split('/')
        location_id = parts[3]
        function_name = parts[5]

        function_data = {
            'project': project_id,
            'location': location_id,
            'name': function_name,
            'status': fnc.status,
            'runtime': fnc.runtime,
            'availableMemoryMb': fnc.availableMemoryMb
        }
        functions.append(function_data)

    return functions

def list_gcs_buckets(project_id):
    def get_bucket_size():
        client = monitoring_v3.MetricServiceClient()

        # Set the time range for the metrics query (last month)
        end_time   = datetime.datetime.utcnow()
        start_time = end_time - datetime.timedelta(days=1)

        interval = monitoring_v3.TimeInterval(
            {
                "start_time":{"seconds": int(start_time.timestamp())},
                "end_time":{"seconds": int(end_time.timestamp())}
            }
        )

        results = client.list_time_series(
            request={
                "name": f'projects/{project_id}',
                "filter": 'metric.type = "storage.googleapis.com/storage/total_bytes"',
                "interval": interval,
                "view": monitoring_v3.ListTimeSeriesRequest.TimeSeriesView.FULL,
            }
        )

        return results
    
    def get_metric_by_bucket(metrics, bucket_name):
        value=0
        for metric in metrics:
            bkt = metric.resource.labels['bucket_name']
            if bkt != bucket_name:
                continue

            if len(metric.points) > 0:
                value=metric.points[0].value.double_value
                break
        return value

    # ------
    # Start
    Logger.log (1, f"PRJ: {project_id} - GCS")

    buckets = []

    metrics_res = get_bucket_size()

    client = discovery.build('storage', 'v1')
    raw_resp = client.buckets().list(project=project_id).execute()
    resp = EasyDict(raw_resp)

    for bkt in resp.items:
        sizeb = get_metric_by_bucket(metrics_res, bkt.name)
        size = GetHumanReadable(sizeb)

        bucket_data = {
            'project': project_id,
            'name': bkt.name,
            'location': bkt.location,
            'storageClass': bkt.storageClass,
            'sizeb': sizeb,
            'size': size
        }
        buckets.append(bucket_data)

    return buckets

def list_gke_clusters(project_id):
    Logger.log (1, f"PRJ: {project_id} - GKE")

    clusters = []
    client = discovery.build('container', 'v1')
    # Note the use of dash (-) to get from all locations
    raw_resp = client.projects().locations().clusters().list(
            parent=f'projects/{project_id}/locations/-'
        ).execute()

    resp = EasyDict(raw_resp)

    if not 'clusters' in resp:
        return clusters

    Logger.log(2, f" |- Found {len(resp.clusters)} GKE clusters")

    for cluster in resp.clusters:
        #print(cluster)

        for np in cluster.nodePools:
            minNodeCount = 0
            maxNodeCount = 0
            initialNodeCount = 0
            if 'autoscaling' in np:
                if 'minNodeCount' in np.autoscaling:
                    minNodeCount = np.autoscaling.minNodeCount
                if 'maxNodeCount' in np.autoscaling:
                    maxNodeCount = np.autoscaling.maxNodeCount
            if 'initialNodeCount' in np:
                initialNodeCount = np.initialNodeCount

            cluster_data = {
                'project': project_id,
                'name': cluster.name,
                'location': cluster.location,
                'nodepool': np.name,
                'currentMasterVersion': cluster.currentMasterVersion,
                'status': cluster.status,
                'totalNodes': cluster.currentNodeCount,
                'machineType': np.config.machineType,
                'diskSizeGb': np.config.diskSizeGb,
                'preemptible': np.get('np.config.preemptible', False),
                'version': np.version,
                'minNodeCount': minNodeCount,
                'maxNodeCount': maxNodeCount,
                'initialNodeCount': initialNodeCount
            }
            clusters.append(cluster_data)

    return clusters

def list_artifact_registry_repos(project_id):
    # https://cloud.google.com/artifact-registry/docs/reference/rest/v1/projects.locations.repositories/list
    Logger.log (1, f"PRJ: {project_id} - Artifact Registry")

    repositories = []

    if not check_api_is_enabled(project_id,'artifactregistry'):
        Logger.log(2, "  | Artifact Registry API is not enabled")
        return repositories
    
    client = discovery.build('artifactregistry', 'v1')

    # get all locations
    raw_resp = client.projects().locations().list(
            name=f'projects/{project_id}'
        ).execute()
    resp_location = EasyDict(raw_resp)

    for location in resp_location.locations:
        Logger.log (3, f'  | Location: {location}')
    
        raw_resp = client.projects().locations().repositories().list(
            parent=f'projects/{project_id}/locations/{location.locationId}'
        ).execute()
        resp = EasyDict(raw_resp)

        # There is not repository in this location
        if not 'repositories' in resp:
            continue

        Logger.log(2, f' |- Found {len(resp.repositories)}')

        for rep in resp.repositories:
            #print(rep)

            sizeB=0
            if 'sizeBytes' in rep:
                sizeB = rep.sizeBytes
            size = GetHumanReadable(sizeb)

            parts=rep.name.split('/')
            name=parts[-1]

            repository_data = {
                'project': project_id,
                'name': name,
                'location': location.locationId,
                'format': rep.format,
                'sizeBytes': sizeB,
                'size': size
            }
            repositories.append(repository_data)

    return repositories

def list_pubsub_topics(project_id):
    Logger.log (1, f"PRJ: {project_id} - PubSub")

    topics = []

    client = discovery.build('pubsub', 'v1')
    raw_resp = client.projects().topics().list(
            project=f'projects/{project_id}'
        ).execute()
    resp = EasyDict(raw_resp)

    if not 'topics' in resp:
        return topics

    Logger.log (2, f' |- Found {len(resp.topics)}')

    for topic in resp.topics:
        parts = topic.name.split('/')
        topic_name = parts[3]

        topic_data = {
            'project': project_id,
            'name': topic_name,
        }
        topics.append(topic_data)

    return topics

def list_networks(project_id):
    Logger.log (1, f"PRJ: {project_id} - Networks")

    networks = {}
    networks['net'] = []
    networks['peer'] = []
    networks['vpn'] = []

    if not check_api_is_enabled(project_id,'compute'):
        Logger.log (2, "  | Compute API is not enabled")
        return networks
    
    # https://cloud.google.com/compute/docs/reference/rest/v1
    client = discovery.build('compute', 'v1')

    # Collect VPN data
    # https://cloud.google.com/compute/docs/reference/rest/v1/vpnTunnels/aggregatedList
    Logger.log(2, 'Searching for VPN')

    r_vpn_agg = client.vpnTunnels().aggregatedList(project=project_id).execute()
    for name, vpn_tunnels_scoped_list in r_vpn_agg['items'].items():
        vpn_resp=EasyDict(vpn_tunnels_scoped_list)

        if 'vpnTunnels' in vpn_resp:
            for vpn in vpn_resp.vpnTunnels:
                Logger.log(3, f'Found VPN {vpn.name}')
                vpn_data = {
                    'project': project_id,
                    'region': vpn.region.split('/')[-1],
                    'name': vpn.name,
                    'status': vpn.status,
                    'description': vpn.description
                }
                networks['vpn'].append(vpn_data)

    # https://cloud.google.com/compute/docs/reference/rest/v1/networks/list
    raw_resp = client.networks().list(project=project_id).execute()
    vpc_resp = EasyDict(raw_resp)

    for vpc in vpc_resp.items:
        network_data = {
            'project': project_id,
            'name': vpc.name
        }
        networks['net'].append(network_data)

        # Collect peering information
        if 'peerings' in vpc:
            for peer in vpc.peerings:
                peer_data = {
                    'project': project_id,
                    'network': vpc.name,
                    'peer_network': peer.network
                }
                networks['peer'].append(peer_data)

    return networks

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-m', '--file', help='File with a list of GCP projects')
    parser.add_argument('-o', '--output', action='append', dest='output', choices=['csv', 'json', 'xls'], default=['csv'], help='Type of output (csv or xls) (can be used multiple times)')
    parser.add_argument('-l', '--log-level', choices=['1', '2', '3'], default='1', help='Log level (1: normal, 2: verbose, 3: debug)')
    parser.add_argument('-r', action='append', dest='options', default=[], help='Specify a resource (compute, sql, functions, gcs, gke, art_repo, pubsub) (can be used multiple times)')
    parser.add_argument('-g', '--gke-ignore', action='store_true', help='Ignore GKE nodes in the compute collect')
    args = parser.parse_args()

    Logger.set_level (int(args.log_level))

    # Check if a file is provided, otherwise use the default project
    if args.file:
        with open(args.file, 'r') as file:
            projects = file.read().splitlines()
    else:
        # If no file is provided, use the current logged-in project
        projects = [get_logged_in_project()]
        Logger.log(2, f"Going to use the already logged-in project {projects}")

    # Create the variables that will be used further
    all_projects = []
    all_instances = []
    all_cloudsql = []
    all_functions = []
    all_gcs_buckets = []
    all_gke_clusters = []
    all_artifact_repos = []
    all_pubsub_topics = []  
    all_networks = {}

    all_projects = list(map(lambda project_id: {'project': project_id}, projects))

    for project in projects:
        Logger.log(1, f"** Processing project: {project} **")

        # Network VPC, VPN, Peerings
        if not args.options or 'network' in args.options:
            networks = list_networks(project)
            all_networks = networks

        # Compute Machines
        if not args.options or 'compute' in args.options:
            instances = list_vm_instances(project, args.gke_ignore)
            all_instances.extend(instances)
        
        # CloudSQL
        if not args.options or 'sql' in args.options:
            cloudsql = list_cloudsql_instances(project)
            all_cloudsql.extend(cloudsql)

        # Functions
        if not args.options or 'functions' in args.options:
            functions = list_functions(project)
            all_functions.extend(functions)

        # GCS Bucket
        if not args.options or 'gcs' in args.options:
            gcs_buckets = list_gcs_buckets(project)
            all_gcs_buckets.extend(gcs_buckets)

        # GKE Clusters
        if not args.options or 'gke' in args.options:
            gke_clusters = list_gke_clusters(project)
            all_gke_clusters.extend(gke_clusters)

        # Artifact Repository
        if not args.options or 'art_repo' in args.options:
            artifact_repos = list_artifact_registry_repos(project)
            all_artifact_repos.extend(artifact_repos)

        # PubSub
        if not args.options or 'pubsub' in args.options:
            pubsub_topics = list_pubsub_topics(project)
            all_pubsub_topics.extend(pubsub_topics)

    Logger.log(1, "Saving to output")

    # Export the collected data to the desired outputs
    Exporter.set_outputs(args.output)
    Exporter.export(all_networks['net'], 'network')
    Exporter.export(all_networks['peer'], 'net_peer')
    Exporter.export(all_networks['vpn'], 'net_vpn')
    Exporter.export(all_instances, 'compute')
    Exporter.export(all_cloudsql, 'cloudsql')
    Exporter.export(all_functions, 'functions')
    Exporter.export(all_gcs_buckets, 'gcs')
    Exporter.export(all_gke_clusters, 'gke')
    Exporter.export(all_artifact_repos, 'artifact')
    Exporter.export(all_pubsub_topics, 'pubsub')
    Exporter.export(all_projects, 'projects')

    Logger.log(1, 'Finished')

if __name__ == '__main__':
    main()